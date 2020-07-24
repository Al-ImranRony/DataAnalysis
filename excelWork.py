import openpyxl
import pandas as pd
import warnings
warnings.filterwarnings("ignore", category=DeprecationWarning)

wb = openpyxl.load_workbook("KPMG_dataset_sprocket_central.xlsx")

#TODO: Get the sheets name of the excel file. 
Sheets = []
Sheets = wb.get_sheet_names()
i = 0
for sheet in Sheets:
    i += 1
    print(str(i)+'.', sheet, end='  ')
print("")

#TODO: Find the number of records, columns & date data in a sheet
Sheet4 = wb.get_sheet_by_name("CustomerDemographic")
print(Sheet4.title, "-", Sheet4.max_row , "records.")
dateCount = 0                                       
for row in Sheet4.rows:                         
    if row[5] is not None:
        dateCount += 1
print(dateCount)

Sheet2 = wb.get_sheet_by_name("Transactions")
dateCount = 0
for row in Sheet2.rows:
    if row[3] is not None:
        dateCount += 1
print(dateCount)

Sheet5 = wb.get_sheet_by_name("CustomerAddress")
print(Sheet5.title, "-", Sheet5.max_row, "records.")
print(Sheet5["A2"].value)
for i in range(2, 8, 2):
        print(i, Sheet5.cell(row=i, column=2).value)

#TODO: Checking if there has an empty cell or not
emptyCnt = 0
for rowNum in range(2, Sheet4.max_row):
    if Sheet4.cell(row=rowNum, column=3).value is None:
        emptyCnt+=1
if emptyCnt > 0:
        print(emptyCnt, "Empty cells in column", 3)

#TODO: Check All Rows and Update Incorrect values
# pvUpdates = {'2': 9, '4': 8, '5': 7}
# for rowNum in range(4, Sheet5.max_row):                  # skip some rows
#     cIds = Sheet5.cell(row=rowNum, column=1).value
#     if cIds in pvUpdates:
#         Sheet5.cell(row=rowNum, column=6).value = pvUpdates[cIds]
# wb.save("updated.xlsx")

wb.create_sheet(index=0, title='tempSheet')
print(wb.get_sheet_names())
wb.remove_sheet(wb.get_sheet_by_name('tempSheet'))
print(wb.get_sheet_names())

for rowOfCellObjects in Sheet5['A2':'C4']:
    for cellObj in rowOfCellObjects:
        print(cellObj.coordinate, cellObj.value)
    print('-------------')